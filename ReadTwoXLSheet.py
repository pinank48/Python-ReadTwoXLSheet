
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook

def ReadSheet1():
    book = load_workbook('items.xlsx')
    #sheet = book.active
    sheet = book['Sheet1']

    rows = sheet.rows
    headers = [cell.value for cell in next(rows)]
    all_rows = []
    for row in rows:
        data = {}
        for title, cell in zip(headers, row):
            data[title] = cell.value
            # print(data)
        all_rows.append(data)
    print(all_rows)

print("#" * 20)

def ReadSheet2():
    book = load_workbook('items.xlsx')
    #sheet = book.active

    sheet2 = book['Sheet2']
    print(sheet2)

    rows2 = sheet2.rows
    headers2 = [cell.value for cell in next(rows2)]
    all_rows2 = []
    for row in rows2:
        data2 = {}
        for title, cell in zip(headers2, row):
            data2[title] = cell.value
            # print(data)
        all_rows2.append(data2)
    print(all_rows2)

root = Tk()
frm = ttk.Frame(root, padding=50)
frm.grid()
ttk.Label(frm, text="Hello World!  ").grid(column=0, row=0)
ttk.Button(frm, text="Sheet-1 ", command=ReadSheet1).grid(column=0, row=1)
ttk.Button(frm, text="Sheet-2 ", command=ReadSheet2).grid(column=0, row=2)
ttk.Button(frm, text="Quit", command=root.destroy).grid(column=0, row=3)
root.mainloop()







