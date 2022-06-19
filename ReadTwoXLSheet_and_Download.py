
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook

def readsheet1():
    book = load_workbook('items.xlsx')
    sheet = book['Sheet1']

    rows = sheet.rows
    headers = [cell.value for cell in next(rows)]
    all_rows = []
    for row in rows:
        data = {}
        for title, cell in zip(headers, row):
            data[title] = cell.value
            # print(data)
        all_rows.append(data)
    print(all_rows)

def readsheet2():
    book = load_workbook('items.xlsx')

    sheet2 = book['Sheet2']
    print(sheet2)

    rows2 = sheet2.rows
    headers2 = [cell.value for cell in next(rows2)]
    all_rows2 = []
    for row in rows2:
        data2 = {}
        for title, cell in zip(headers2, row):
            data2[title] = cell.value
            # print(data)
        all_rows2.append(data2)
    print(all_rows2)

    #book.save("demo.xlsx")

def download():
    book = load_workbook('items.xlsx')
    book.save("download.xlsx")
    print("Download Sucessfully!")

root = Tk()
frm = ttk.Frame(root, padding=50)
frm.grid()
ttk.Label(frm, text="Hello World!  ").grid(column=0, row=0)
ttk.Button(frm, text="Sheet-1 ", command=readsheet1).grid(column=0, row=1)
ttk.Button(frm, text="Sheet-2 ", command=readsheet2).grid(column=0, row=2)
ttk.Button(frm, text="Download", command=download).grid(column=0, row=3)
ttk.Button(frm, text="Quit", command=root.destroy).grid(column=0, row=4)
root.mainloop()

'''
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook

def GetXLDat() :
        book = load_workbook('items.xlsx')
        sheet = book.active

        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []
        for row in rows:
            data = {}
            for title, cell in zip(headers, row):
                data[title] = cell.value
                # print(data)
            all_rows.append(data)
        print(all_rows)


root = Tk()
frm = ttk.Frame(root, padding=30)
root.title('Self Calculator')
frm.grid()
ttk.Label(frm, text="Hello World!").grid(column=0, row=0)
ttk.Button(frm, text="Quit", command=GetXLDat()).grid(column=1, row=0)
root.mainloop()
'''